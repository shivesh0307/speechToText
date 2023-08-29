import os.path
import tkinter as tk
import threading
import speech_recognition as sr
from openpyxl import Workbook
from openpyxl import load_workbook
from textblob import TextBlob

class SpeechToTextApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Speech to Text and Sentiment Analysis")
        
        self.record_button = tk.Button(root, text="Record", command=self.start_recording)
        self.record_button.pack(pady=10)
        
        self.text_display = tk.Text(root, height=10, width=40)
        self.text_display.pack(padx=10, pady=10)
        
        self.recognizer = sr.Recognizer()
        self.recording_thread = None
        self.output_file = "speech_analysis.xlsx"  # Excel file to store speech analysis
        
        self.workbook = None
        self.sheet = None
        self.create_or_load_excel_file()
        
    def create_or_load_excel_file(self):
        if os.path.isfile(self.output_file):
            self.workbook = load_workbook(self.output_file)
            self.sheet = self.workbook.active
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(["Recognized Text", "Sentiment Score"])
        
    def start_recording(self):
        self.record_button.config(state="disabled")
        self.text_display.delete(1.0, tk.END)
        self.text_display.insert(tk.END, "Recording...")
        
        self.recording_thread = threading.Thread(target=self.process_audio)
        self.recording_thread.start()
        
    def process_audio(self):
        with sr.Microphone() as source:
            audio = self.recognizer.listen(source)
            
        try:
            text = self.recognizer.recognize_google(audio)
            self.text_display.delete(1.0, tk.END)
            self.text_display.insert(tk.END, text)
            sentiment_score = self.perform_sentiment_analysis(text)
            self.store_text_to_excel(text, sentiment_score)  # Store recognized text and sentiment score
            self.display_sentiment_score(sentiment_score)
        except sr.UnknownValueError:
            self.text_display.delete(1.0, tk.END)
            self.text_display.insert(tk.END, "Could not understand audio")
        except sr.RequestError as e:
            self.text_display.delete(1.0, tk.END)
            self.text_display.insert(tk.END, f"Could not request results; {e}")
            
        self.record_button.config(state="normal")
        
    def perform_sentiment_analysis(self, text):
        blob = TextBlob(text)
        sentiment_score = blob.sentiment.polarity
        return sentiment_score
        
    def store_text_to_excel(self, text, sentiment_score):
        self.sheet.append([text, sentiment_score])
        
    def display_sentiment_score(self, sentiment_score):
        scaled_score = sentiment_score * 10
        sentiment_label = ""
        if scaled_score > 5:
            sentiment_label = "Positive"
        elif scaled_score < 5:
            sentiment_label = "Negative"
        else:
            sentiment_label = "Neutral"
        
        self.text_display.insert(tk.END, f"\nSentiment Score: {scaled_score:.2f} ({sentiment_label})\n\n")
        
    def save_excel_file(self):
        self.workbook.save(self.output_file)
        
if __name__ == "__main__":
    root = tk.Tk()
    app = SpeechToTextApp(root)
    root.mainloop()
    app.save_excel_file()  # Save Excel file before closing the application
